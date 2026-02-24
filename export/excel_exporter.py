"""Excel export functionality."""

import pandas as pd
from datetime import date, time, timedelta
from typing import List, Dict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from core.models import WeekSchedule, DaySchedule, Assignment, EmployeeAvailability
from core.constants import CATEGORIES, CATEGORY_COVERAGE
from core.utils import generate_time_blocks, is_weekend
from engine.weekend_tracker import WeekendTracker


class ExcelExporter:
    """Exports schedule to Excel format with 3 tabs."""

    def __init__(self, weekend_tracker: WeekendTracker):
        self.weekend_tracker = weekend_tracker

    def export_schedule(
        self,
        week_schedule: WeekSchedule,
        output_path: str
    ) -> str:
        """
        Export week schedule to Excel.

        Args:
            week_schedule: Schedule to export
            output_path: Path to save Excel file

        Returns:
            Path to created file
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Tab 1: WorkHours
        self._create_workhours_tab(wb, week_schedule)

        # Tab 2: WeekdaysGrid
        self._create_weekdays_grid_tab(wb, week_schedule)

        # Tab 3: WeekendGrid
        self._create_weekend_grid_tab(wb, week_schedule)

        wb.save(output_path)
        return output_path

    def _create_workhours_tab(self, wb: Workbook, week_schedule: WeekSchedule):
        """Create WorkHours tab."""
        ws = wb.create_sheet("WorkHours")

        # Headers
        headers = ["Date", "DayOfWeek", "Employee", "Status", "StartTime", "EndTime", "Notes"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data rows
        for day_schedule in week_schedule.days:
            for avail in day_schedule.availability:
                # Get actual start/end times
                from core.constants import EMPLOYEE_DEFAULT_HOURS
                default_start, default_end = EMPLOYEE_DEFAULT_HOURS.get(
                    avail.employee_name,
                    (time(9, 0), time(17, 0))
                )

                start_time = avail.start_time or default_start
                end_time = avail.end_time or default_end

                row = [
                    day_schedule.date.strftime("%Y-%m-%d"),
                    day_schedule.day_of_week,
                    avail.employee_name,
                    avail.status,
                    start_time.strftime("%H:%M"),
                    end_time.strftime("%H:%M"),
                    avail.notes
                ]
                ws.append(row)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

    def _create_weekdays_grid_tab(self, wb: Workbook, week_schedule: WeekSchedule):
        """Create WeekdaysGrid tab."""
        ws = wb.create_sheet("WeekdaysGrid")

        # Get weekday schedules only
        weekday_schedules = [
            day for day in week_schedule.days
            if not is_weekend(day.date)
        ]

        # Generate time blocks
        time_blocks = generate_time_blocks(8, 24, 30)

        # Create grid for each weekday
        for day_schedule in weekday_schedules:
            # Day header
            ws.append([f"{day_schedule.day_of_week} - {day_schedule.date.strftime('%Y-%m-%d')}"])
            ws[ws.max_row][0].font = Font(bold=True, size=12)

            # Column headers: Time, Categories...
            headers = ["Time"] + CATEGORIES
            ws.append(headers)

            # Style header row
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Data rows
            for block in time_blocks:
                row = [str(block)]

                for category in CATEGORIES:
                    # Find assignments for this category and block
                    assignments = [
                        a for a in day_schedule.assignments
                        if a.category == category and a.time_block == block
                    ]

                    if assignments:
                        employees = ", ".join(sorted(set(a.employee_name for a in assignments)))
                        row.append(employees)
                    else:
                        row.append("")

                ws.append(row)

            # Coverage summary
            ws.append([])
            ws.append(["Coverage Summary"])
            ws[ws.max_row][0].font = Font(bold=True)

            for category in CATEGORIES:
                cat_coverage = self._calculate_coverage(day_schedule, category)
                ws.append([category, f"{cat_coverage:.1f}%"])

            # Uncovered blocks
            uncovered = self._find_uncovered_blocks(day_schedule)
            if uncovered:
                ws.append([])
                ws.append(["Uncovered Blocks"])
                ws[ws.max_row][0].font = Font(bold=True, color="FF0000")

                for cat, blocks in uncovered.items():
                    ws.append([cat, ", ".join(str(b) for b in blocks)])

            ws.append([])
            ws.append([])

    def _create_weekend_grid_tab(self, wb: Workbook, week_schedule: WeekSchedule):
        """Create WeekendGrid tab."""
        ws = wb.create_sheet("WeekendGrid")

        # Get weekend schedules only
        weekend_schedules = [
            day for day in week_schedule.days
            if is_weekend(day.date)
        ]

        # Generate time blocks
        time_blocks = generate_time_blocks(8, 24, 30)

        # Create grid for each weekend day
        for day_schedule in weekend_schedules:
            # Day header
            ws.append([f"{day_schedule.day_of_week} - {day_schedule.date.strftime('%Y-%m-%d')}"])
            ws[ws.max_row][0].font = Font(bold=True, size=12)

            # Column headers
            headers = ["Time"] + CATEGORIES
            ws.append(headers)

            # Style header row
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Data rows
            for block in time_blocks:
                row = [str(block)]

                for category in CATEGORIES:
                    assignments = [
                        a for a in day_schedule.assignments
                        if a.category == category and a.time_block == block
                    ]

                    if assignments:
                        employees = ", ".join(sorted(set(a.employee_name for a in assignments)))
                        row.append(employees)
                    else:
                        row.append("")

                ws.append(row)

            ws.append([])
            ws.append([])

        # Weekend tracking summary
        ws.append(["Weekend Work Summary"])
        ws[ws.max_row][0].font = Font(bold=True, size=14)
        ws.append([])

        # Get month from week start
        month = week_schedule.week_start
        summary = self.weekend_tracker.get_weekend_summary(month)

        # Headers
        ws.append(["Employee", "Weekends Off", "Worked Saturday", "Worked Sunday", "Total Worked", "Compliant"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        # Data
        for employee, tracking in sorted(summary.items()):
            compliant = "Yes" if tracking.is_compliant else "No"
            row = [
                employee,
                tracking.weekends_off,
                tracking.weekends_worked_saturday,
                tracking.weekends_worked_sunday,
                tracking.total_weekends_worked,
                compliant
            ]
            ws.append(row)

            # Color non-compliant rows
            if not tracking.is_compliant:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def _calculate_coverage(self, day_schedule: DaySchedule, category: str) -> float:
        """Calculate coverage percentage for a category."""
        from core.constants import CATEGORY_COVERAGE

        if category not in CATEGORY_COVERAGE:
            return 0.0

        coverage_start, coverage_end = CATEGORY_COVERAGE[category]

        # Generate required blocks
        required_blocks = generate_time_blocks(8, 24, 30)

        # Filter to coverage window
        relevant_blocks = [
            b for b in required_blocks
            if self._time_in_range(b.start_time, coverage_start, coverage_end)
        ]

        if not relevant_blocks:
            return 0.0

        # Count covered blocks
        covered = 0
        for block in relevant_blocks:
            assignments = [
                a for a in day_schedule.assignments
                if a.category == category and a.time_block == block
            ]
            if assignments:
                covered += 1

        return (covered / len(relevant_blocks)) * 100.0

    def _time_in_range(self, check_time: time, start: time, end: time) -> bool:
        """Check if time is in range, handling midnight crossing."""
        if end < start:  # Crosses midnight
            return check_time >= start or check_time < end
        else:
            return start <= check_time < end

    def _find_uncovered_blocks(self, day_schedule: DaySchedule) -> Dict[str, List[str]]:
        """Find uncovered time blocks per category."""
        uncovered = {}

        for category in CATEGORIES:
            from core.constants import CATEGORY_COVERAGE
            coverage_start, coverage_end = CATEGORY_COVERAGE[category]

            required_blocks = generate_time_blocks(8, 24, 30)
            relevant_blocks = [
                b for b in required_blocks
                if self._time_in_range(b.start_time, coverage_start, coverage_end)
            ]

            missing = []
            for block in relevant_blocks:
                assignments = [
                    a for a in day_schedule.assignments
                    if a.category == category and a.time_block == block
                ]
                if not assignments:
                    missing.append(str(block))

            if missing:
                uncovered[category] = missing

        return uncovered
